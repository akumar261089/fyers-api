"""
main_get_zones.py
-----------------
Step 1 – enrich raw CSVs with technical indicators
Step 2 – scan enriched CSVs for Supply/Demand zones (GEL pattern)
Step 3 – save a consolidated report (CSV + Excel)

Zone Logic (from notes):
  C1 = Leg-In  candle → TR > ATR  (filled candle)
  C2 = UOC     candle → TR < ATR  (boring candle, dual-side wicks)
  C3 = Leg-Out candle → TR > ATR  (filled candle, strong move)

  zone_high = max(C2.open, C2.close)
  zone_low  = min(C2.open, C2.close)

  Supply Zone → C3 leg-out is DOWN  (price returns to zone to SELL)
  Demand Zone → C3 leg-out is UP    (price returns to zone to BUY)
"""

import logging
from pathlib import Path

from technical_indicators import add_indicators_to_folder
from pattern_detector      import scan_patterns

# ── logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)

# ═════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION  –  edit these paths / settings
# ═════════════════════════════════════════════════════════════════════════════

RAW_FOLDER      = "data/raw"        # your input CSVs
ENRICHED_FOLDER = "data/enriched"   # indicator-enriched CSVs (auto-created)
REPORT_FOLDER   = "data/reports"    # where the final report is saved

# Indicator parameters
ATR_PERIOD = 14
SMA_PERIOD = 20
EMA_PERIOD = 20
RSI_PERIOD = 14
BB_PERIOD  = 20
BB_STD     = 2.0

# Pattern parameters
# C1 body must be > BODY_MULTIPLIER × C2 body
# C3 body must be > BODY_MULTIPLIER × C2 body
BODY_MULTIPLIER = 3
ATR_COLUMN      = f"atr_{ATR_PERIOD}"

# Report output formats  (set False to skip)
SAVE_CSV   = True
SAVE_EXCEL = True


# ═════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═════════════════════════════════════════════════════════════════════════════

def main() -> None:

    # ── Step 1: enrich ────────────────────────────────────────────────────────
    print("\n" + "═" * 60)
    print("  STEP 1  –  Adding technical indicators")
    print("═" * 60)

    add_indicators_to_folder(
        input_folder  = RAW_FOLDER,
        output_folder = ENRICHED_FOLDER,
        atr_period    = ATR_PERIOD,
        sma_period    = SMA_PERIOD,
        ema_period    = EMA_PERIOD,
        rsi_period    = RSI_PERIOD,
        bb_period     = BB_PERIOD,
        bb_std        = BB_STD,
    )

    # ── Step 2: scan zones ────────────────────────────────────────────────────
    print("\n" + "═" * 60)
    print("  STEP 2  –  Scanning for Supply / Demand Zones (GEL pattern)")
    print("═" * 60)

    report = scan_patterns(
        enriched_folder = ENRICHED_FOLDER,
        atr_col         = ATR_COLUMN,
        body_multiplier = BODY_MULTIPLIER,
    )

    if report.empty:
        print("No zones found – skipping sorting.")
    else:
        required_cols = ["strength", "score", "c3_datetime"]
        missing_cols  = [col for col in required_cols if col not in report.columns]

        if missing_cols:
            print(f"Warning: Missing columns {missing_cols} – skipping sort.")
        else:
            report = report.sort_values(
                required_cols,
                ascending=[False, False, True]
            )

    # ── Step 3: save report ───────────────────────────────────────────────────
    print("\n" + "═" * 60)
    print("  STEP 3  –  Saving report")
    print("═" * 60)

    report_dir = Path(REPORT_FOLDER)
    report_dir.mkdir(parents=True, exist_ok=True)

    if report.empty:
        print("  No zones found – report not written.")
        return

    if SAVE_CSV:
        csv_path = report_dir / "zone_report.csv"
        report.to_csv(csv_path, index=False)
        print(f"  CSV   saved → {csv_path}  ({len(report)} rows)")

    if SAVE_EXCEL:
        try:
            import openpyxl                          # noqa: F401
            xl_path = report_dir / "zone_report.xlsx"
            _save_excel(report, xl_path)
            print(f"  Excel saved → {xl_path}")
        except ImportError:
            print("  [!] openpyxl not installed – Excel output skipped.")
            print("      pip install openpyxl")

    # ── print summary ─────────────────────────────────────────────────────────
    print("\n" + "─" * 60)
    print(f"  Total zones found : {len(report)}")
    for ptype, n in report["pattern_label"].value_counts().items():
        print(f"    {ptype:<30} {n}")
    print("\n  Zone columns in report:")
    print("    zone_high  = max(C2.open, C2.close)  ← ZONE TOP")
    print("    zone_low   = min(C2.open, C2.close)  ← ZONE BOTTOM")
    print("    zone_type  = Supply | Demand")
    print("    sub_type   = DBD | RBD | DBR | RBR")
    print("─" * 60)


# ─────────────────────────────────────────────────────────────────────────────
#  Excel formatter
# ─────────────────────────────────────────────────────────────────────────────

def _save_excel(report, path: Path) -> None:
    """Write report to Excel with conditional row colouring."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # Demand zone = green (price returns to BUY)
    # Supply zone = red   (price returns to SELL)
    GREEN_FILL = PatternFill("solid", fgColor="CCFFCC")
    RED_FILL   = PatternFill("solid", fgColor="FFCCCC")
    HDR_FILL   = PatternFill("solid", fgColor="1F3864")
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
    BODY_FONT  = Font(size=10)

    wb = openpyxl.Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    counts = report["pattern_label"].value_counts()
    rows = [
        ["Supply / Demand Zone Report  (GEL Pattern)"],
        [],
        ["Total zones found",     len(report)],
    ]
    for pt, n in counts.items():
        rows.append([f"  {pt}", n])
    rows += [
        [],
        ["Symbols with zones",    report["symbol"].nunique()],
        ["Files scanned",         report["source_file"].nunique()],
        [],
        ["── Zone Logic ──"],
        ["C1  Leg-In  candle", "TR > ATR  (filled candle)"],
        ["C2  UOC     candle", "TR < ATR  (boring, dual-side wicks)  ← ZONE CANDLE"],
        ["C3  Leg-Out candle", "TR > ATR  (filled candle, strong move)"],
        [],
        ["zone_high", "max(C2.open, C2.close)  ← zone top"],
        ["zone_low",  "min(C2.open, C2.close)  ← zone bottom"],
        [],
        ["Supply Zone", "C3 leg-out is DOWN  → return to zone to SELL"],
        ["Demand Zone", "C3 leg-out is UP    → return to zone to BUY"],
        [],
        ["Sub-types"],
        ["DBD", "Drop-Base-Drop  (Supply)"],
        ["RBD", "Rally-Base-Drop (Supply)"],
        ["DBR", "Drop-Base-Rally (Demand)"],
        ["RBR", "Rally-Base-Rally (Demand)"],
        [],
        ["Row colour key"],
        ["Green row", "Demand Zone  (return to BUY)"],
        ["Red   row", "Supply Zone  (return to SELL)"],
    ]
    for row in rows:
        ws_sum.append(row)

    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum.column_dimensions["A"].width = 34
    ws_sum.column_dimensions["B"].width = 58

    # ── Detail sheet ──────────────────────────────────────────────────────────
    def _write_sheet(ws, data: "pd.DataFrame") -> None:
        headers = list(data.columns)
        ws.append(headers)
        for cell in ws[1]:
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = Alignment(horizontal="center")

        _FILL_MAP = {
            "breakout_up":   GREEN_FILL,   # Demand zone
            "breakout_down": RED_FILL,     # Supply zone
        }

        pt_col = headers.index("pattern_type") + 1

        for _, row_data in data.iterrows():
            ws.append([
                v if isinstance(v, (int, float)) else str(v)
                for v in row_data
            ])
            rn   = ws.max_row
            fill = _FILL_MAP.get(ws.cell(rn, pt_col).value, RED_FILL)
            for ci in range(1, len(headers) + 1):
                c = ws.cell(rn, ci)
                c.fill = fill
                c.font = BODY_FONT

        # auto column width
        for ci, hdr in enumerate(headers, 1):
            col_letter = get_column_letter(ci)
            max_len = max(
                len(str(hdr)),
                *(len(str(ws.cell(r, ci).value or ""))
                  for r in range(2, ws.max_row + 1)),
                8,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

        ws.freeze_panes = "A2"

    ws_all = wb.create_sheet("All Zones")
    _write_sheet(ws_all, report)

    # ── Per-symbol sheets ─────────────────────────────────────────────────────
    for symbol, grp in report.groupby("symbol"):
        name = str(symbol).replace("NSE:", "").replace("-EQ", "")[:31]
        _write_sheet(wb.create_sheet(name), grp.reset_index(drop=True))

    wb.save(path)


if __name__ == "__main__":
    main()