"""
main.py
-------
Step 1 – enrich raw CSVs with technical indicators
Step 2 – scan enriched CSVs for the NR-Inside-Expansion pattern
Step 3 – save a consolidated report (CSV + Excel)
"""

import logging
from pathlib import Path

from technical_indicators import add_indicators_to_folder
from pattern_detector      import scan_patterns

# ── logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)

# ═════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION  –  edit these paths / settings
# ═════════════════════════════════════════════════════════════════════════════

RAW_FOLDER      = "data/raw"        # your input CSVs
ENRICHED_FOLDER = "data/enriched"   # indicator-enriched CSVs (auto-created)
REPORT_FOLDER   = "data/reports"    # where the final report is saved

# Indicator parameters
ATR_PERIOD = 14
SMA_PERIOD = 20
EMA_PERIOD = 20
RSI_PERIOD = 14
BB_PERIOD  = 20
BB_STD     = 2.0

# Pattern parameters
BODY_MULTIPLIER = 1.2               # C1 > 2×C2  and  C3 > 2×C1
ATR_COLUMN      = f"atr_{ATR_PERIOD}"

# Report output formats  (set False to skip)
SAVE_CSV   = True
SAVE_EXCEL = True


# ═════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═════════════════════════════════════════════════════════════════════════════

def main() -> None:

    # ── Step 1: enrich ────────────────────────────────────────────────────────
    print("\n" + "═" * 60)
    print("  STEP 1  –  Adding technical indicators")
    print("═" * 60)

    add_indicators_to_folder(
        input_folder  = RAW_FOLDER,
        output_folder = ENRICHED_FOLDER,
        atr_period    = ATR_PERIOD,
        sma_period    = SMA_PERIOD,
        ema_period    = EMA_PERIOD,
        rsi_period    = RSI_PERIOD,
        bb_period     = BB_PERIOD,
        bb_std        = BB_STD,
    )

    # ── Step 2: scan patterns ─────────────────────────────────────────────────
    print("\n" + "═" * 60)
    print("  STEP 2  –  Scanning for NR-Inside-Expansion patterns")
    print("═" * 60)

    report = scan_patterns(
        enriched_folder = ENRICHED_FOLDER,
        atr_col         = ATR_COLUMN,
        body_multiplier = BODY_MULTIPLIER,
    )

    # ✅ FIX: handle empty or missing columns
    if report.empty:
        print("No patterns found – skipping sorting.")
    else:
        required_cols = ["strength", "score", "c3_datetime"]

        missing_cols = [col for col in required_cols if col not in report.columns]

        if missing_cols:
            print(f"Warning: Missing columns {missing_cols} – skipping sort.")
        else:
            report = report.sort_values(
                required_cols,
                ascending=[False, False, True]
            )
    # ── Step 3: save report ───────────────────────────────────────────────────
    print("\n" + "═" * 60)
    print("  STEP 3  –  Saving report")
    print("═" * 60)

    report_dir = Path(REPORT_FOLDER)
    report_dir.mkdir(parents=True, exist_ok=True)

    if report.empty:
        print("  No patterns found – report not written.")
        return

    if SAVE_CSV:
        csv_path = report_dir / "pattern_report.csv"
        report.to_csv(csv_path, index=False)
        print(f"  CSV   saved → {csv_path}  ({len(report)} rows)")

    if SAVE_EXCEL:
        try:
            import openpyxl                          # noqa: F401
            xl_path = report_dir / "pattern_report.xlsx"
            _save_excel(report, xl_path)
            print(f"  Excel saved → {xl_path}")
        except ImportError:
            print("  [!] openpyxl not installed – Excel output skipped.")
            print("      pip install openpyxl")

    # ── print summary ─────────────────────────────────────────────────────────
    print("\n" + "─" * 60)
    print(f"  Total patterns  : {len(report)}")
    for ptype, n in report["pattern_type"].value_counts().items():
        print(f"    {ptype:<30} {n}")
    print("─" * 60)


# ─────────────────────────────────────────────────────────────────────────────
#  Excel formatter
# ─────────────────────────────────────────────────────────────────────────────

def _save_excel(report, path: Path) -> None:
    """Write report to Excel with conditional row colouring."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    GREEN_FILL  = PatternFill("solid", fgColor="CCFFCC")   # breakout_up
    RED_FILL    = PatternFill("solid", fgColor="FFCCCC")   # breakout_down
    HDR_FILL   = PatternFill("solid", fgColor="1F3864")
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
    BODY_FONT  = Font(size=10)

    wb = openpyxl.Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    counts = report["pattern_type"].value_counts()
    rows = [
        ["NR-Inside-Expansion Pattern Report"],
        [],
        ["Total patterns found",     len(report)],
    ]
    for pt, n in counts.items():
        rows.append([f"  {pt}", n])
    rows += [
        [],
        ["Symbols with matches",   report["symbol"].nunique()],
        ["Files scanned",          report["source_file"].nunique()],
        [],
        ["── Pattern Logic ──"],
        ["Condition 1",  "C1 TR > ATR  |  C2 TR < ATR  |  C3 TR > ATR"],
        ["Condition 2",  f"C1 body > {BODY_MULTIPLIER}× C2 body  AND  "
                         f"C3 body > {BODY_MULTIPLIER}× C1 body"],
        ["Condition 3",  "C3 body must clear C2 body entirely (no overlap)"],
        ["  breakout_up",   "C3 open > C2 top  AND  C3 close > C2 top"],
        ["  breakout_down", "C3 open < C2 bot  AND  C3 close < C2 bot"],
        [],
        ["Row colour key"],
        ["Green row", "breakout_up   (C3 body entirely above C2 body)"],
        ["Red   row", "breakout_down (C3 body entirely below C2 body)"],
    ]
    for row in rows:
        ws_sum.append(row)

    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum.column_dimensions["A"].width = 34
    ws_sum.column_dimensions["B"].width = 58

    # ── Detail sheet ──────────────────────────────────────────────────────────
    def _write_sheet(ws, data: "pd.DataFrame") -> None:
        headers = list(data.columns)
        ws.append(headers)
        for cell in ws[1]:
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = Alignment(horizontal="center")

        _FILL_MAP = {
            "breakout_up":   GREEN_FILL,
            "breakout_down": RED_FILL,
        }

        pt_col = headers.index("pattern_type") + 1

        for _, row_data in data.iterrows():
            ws.append([
                v if isinstance(v, (int, float)) else str(v)
                for v in row_data
            ])
            rn   = ws.max_row
            fill = _FILL_MAP.get(ws.cell(rn, pt_col).value, RED_FILL)
            for ci in range(1, len(headers) + 1):
                c = ws.cell(rn, ci)
                c.fill = fill
                c.font = BODY_FONT

        # auto column width
        for ci, hdr in enumerate(headers, 1):
            col_letter = get_column_letter(ci)
            max_len = max(
                len(str(hdr)),
                *(len(str(ws.cell(r, ci).value or ""))
                  for r in range(2, ws.max_row + 1)),
                8,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

        ws.freeze_panes = "A2"

    ws_all = wb.create_sheet("All Patterns")
    _write_sheet(ws_all, report)

    # ── Per-symbol sheets ─────────────────────────────────────────────────────
    for symbol, grp in report.groupby("symbol"):
        name = str(symbol).replace("NSE:", "").replace("-EQ", "")[:31]
        _write_sheet(wb.create_sheet(name), grp.reset_index(drop=True))

    wb.save(path)


if __name__ == "__main__":
    main()
